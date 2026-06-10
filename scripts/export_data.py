import urllib.request
import json
import csv
import sys
import os
from datetime import datetime

# Database credentials
KV_REST_API_URL = "https://touched-gnu-145054.upstash.io"
KV_REST_API_TOKEN = "gQAAAAAAAjaeAAIgcDE1YmFiYzUwNGI2OTM0ZjY0OTcxZmFiZmJmM2IyNzMxZA"

def fetch_data():
    url = f"{KV_REST_API_URL}/get/election_data?_token={KV_REST_API_TOKEN}"
    print(f"Connecting to Upstash Redis to fetch election data...")
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req) as response:
            if response.status == 200:
                raw = json.loads(response.read().decode('utf-8'))
                result = raw.get("result")
                if isinstance(result, str):
                    return json.loads(result)
                return result
    except Exception as e:
        print(f"Error fetching from Redis API: {e}", file=sys.stderr)
    return None

def align_series_to_history(history, series):
    """
    Given the list of global updates (history) and a region's series,
    reconstruct the state (votes, pct) of the region for EACH history entry.
    Since both are chronological, we align them.
    """
    aligned = []
    series_idx = 0
    current_state = {
        "kv": 0,
        "rv": 0,
        "a": 50.0,
        "b": 50.0
    }
    
    # We find the first series point that has kv/rv to initialize if possible
    first_with_votes = None
    for pt in series:
        if "kv" in pt and "rv" in pt:
            first_with_votes = pt
            break
            
    if first_with_votes:
        current_state["kv"] = first_with_votes["kv"]
        current_state["rv"] = first_with_votes["rv"]
        current_state["a"] = first_with_votes["a"]
        current_state["b"] = first_with_votes["b"]

    for hist_entry in history:
        hist_time = hist_entry.get("time_display")
        
        # Check if there is a series point matching this time_display
        # Advance series_idx accordingly.
        matched_pt = None
        temp_idx = series_idx
        while temp_idx < len(series):
            pt = series[temp_idx]
            if pt.get("t") == hist_time:
                matched_pt = pt
                series_idx = temp_idx
                break
            temp_idx += 1
            
        if matched_pt:
            if "kv" in matched_pt and "rv" in matched_pt:
                current_state["kv"] = matched_pt["kv"]
                current_state["rv"] = matched_pt["rv"]
            current_state["a"] = matched_pt.get("a", current_state["a"])
            current_state["b"] = matched_pt.get("b", current_state["b"])
        
        aligned.append({
            "keiko_votos": current_state["kv"],
            "roberto_votos": current_state["rv"],
            "keiko_pct": current_state["b"],
            "roberto_pct": current_state["a"],
            "validos": current_state["kv"] + current_state["rv"]
        })
        
    return aligned

def main():
    # Setup directory paths relative to this script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.abspath(os.path.join(script_dir, ".."))
    reports_dir = os.path.join(root_dir, "reports")
    
    # Ensure reports directory exists
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir, exist_ok=True)
        
    csv_file = os.path.join(reports_dir, "reporte_ultimas_20_actualizaciones.csv")
    xls_file = os.path.join(reports_dir, "reporte_ultimas_20_actualizaciones.xls")
    xlsx_file = os.path.join(reports_dir, "reporte_ultimas_20_actualizaciones.xlsx")

    data = fetch_data()
    if not data:
        # Fallback to local data.json in root
        data_json_path = os.path.join(root_dir, "data.json")
        print(f"API fetch failed. Falling back to local data at {data_json_path}...")
        if os.path.exists(data_json_path):
            with open(data_json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        else:
            print("Error: Local data.json not found and API call failed.", file=sys.stderr)
            sys.exit(1)
            
    history = data.get("projections_history", [])
    regiones = data.get("regiones", [])
    extranjero = data.get("extranjero", [])
    
    if not history:
        print("Error: No history found in data.", file=sys.stderr)
        sys.exit(1)
        
    total_updates = len(history)
    print(f"Total updates in database: {total_updates}")
    
    # We want the last 20 updates
    num_updates_to_export = min(20, total_updates)
    last_20_history = history[-num_updates_to_export:]
    
    # Align each region's series to the global history
    region_aligned_data = {}
    for r in regiones:
        ubigeo = r.get("ubigeo")
        series = r.get("serie", [])
        aligned = align_series_to_history(history, series)
        region_aligned_data[ubigeo] = aligned[-num_updates_to_export:]
        
    # Align each continent's series to the global history
    continent_aligned_data = {}
    for c in extranjero:
        ubigeo = c.get("ubigeo")
        series = c.get("serie", [])
        aligned = align_series_to_history(history, series)
        continent_aligned_data[ubigeo] = aligned[-num_updates_to_export:]
        
    # Headers for our flat CSV file
    headers = [
        "Nro Actualización",
        "Fecha y Hora (Lima)",
        "Ámbito",
        "Ubigeo",
        "Nombre Región/Continente",
        "Votos Keiko",
        "Votos Roberto",
        "Votos Válidos",
        "Keiko %",
        "Roberto %",
        "Nacional Actas Perú %",
        "Nacional Actas Extranjero %",
        "Nacional Keiko Votos",
        "Nacional Roberto Votos",
        "Nacional Keiko %",
        "Nacional Roberto %"
    ]
    
    csv_rows = []
    
    # Loop through the last 20 updates
    for u_idx, hist_entry in enumerate(last_20_history):
        update_num = total_updates - num_updates_to_export + u_idx + 1
        timestamp = hist_entry.get("timestamp", "")
        time_display = hist_entry.get("time_display", "")
        
        # format date-time nicely from timestamp (e.g. 2026-06-10T11:18:00-05:00)
        try:
            dt = datetime.fromisoformat(timestamp)
            formatted_dt = dt.strftime("%d/%m/%Y %I:%M:%S %p")
        except:
            formatted_dt = f"{timestamp} ({time_display})"
            
        # Add Regions
        for r_idx, r in enumerate(regiones):
            ubigeo = r.get("ubigeo")
            name = r.get("nombre")
            aligned = region_aligned_data[ubigeo][u_idx]
            
            csv_rows.append([
                update_num,
                formatted_dt,
                "Región",
                ubigeo,
                name,
                aligned["keiko_votos"],
                aligned["roberto_votos"],
                aligned["validos"],
                f"{aligned['keiko_pct']:.3f}%" if isinstance(aligned['keiko_pct'], (int, float)) else aligned['keiko_pct'],
                f"{aligned['roberto_pct']:.3f}%" if isinstance(aligned['roberto_pct'], (int, float)) else aligned['roberto_pct'],
                f"{hist_entry.get('acts_pct', 0):.3f}%",
                f"{hist_entry.get('ex_acts_pct', 0):.3f}%",
                hist_entry.get("current_keiko", 0),
                hist_entry.get("current_roberto", 0),
                f"{hist_entry.get('current_keiko_pct', 50.0):.3f}%",
                f"{hist_entry.get('current_roberto_pct', 50.0):.3f}%"
            ])
            
        # Add Continents
        for c_idx, c in enumerate(extranjero):
            ubigeo = c.get("ubigeo")
            name = c.get("nombre")
            aligned = continent_aligned_data[ubigeo][u_idx]
            
            csv_rows.append([
                update_num,
                formatted_dt,
                "Extranjero",
                ubigeo,
                name,
                aligned["keiko_votos"],
                aligned["roberto_votos"],
                aligned["validos"],
                f"{aligned['keiko_pct']:.3f}%" if isinstance(aligned['keiko_pct'], (int, float)) else aligned['keiko_pct'],
                f"{aligned['roberto_pct']:.3f}%" if isinstance(aligned['roberto_pct'], (int, float)) else aligned['roberto_pct'],
                f"{hist_entry.get('acts_pct', 0):.3f}%",
                f"{hist_entry.get('ex_acts_pct', 0):.3f}%",
                hist_entry.get("current_keiko", 0),
                hist_entry.get("current_roberto", 0),
                f"{hist_entry.get('current_keiko_pct', 50.0):.3f}%",
                f"{hist_entry.get('current_roberto_pct', 50.0):.3f}%"
            ])
            
    # Write to CSV
    with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(headers)
        writer.writerows(csv_rows)
    print(f"Successfully generated CSV file: {csv_file}")
    
    # Generate styled XLS file (XML/HTML format)
    generate_styled_xls(xls_file, headers, csv_rows)
    print(f"Successfully generated styled XLS file: {xls_file}")
    
    # Check if openpyxl is installed and generate a true .xlsx file
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Actualizaciones"
        
        # Styles
        title_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
        data_font = Font(name="Calibri", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        border_thin = Side(border_style="thin", color="D3D3D3")
        cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = center_align
            
        # Write rows
        for row_idx, row_data in enumerate(csv_rows, 2):
            # Zebra striping
            bg_color = "F2F5F8" if row_idx % 2 == 0 else "FFFFFF"
            row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.fill = row_fill
                cell.border = cell_border
                
                # Alignments and number formats
                if col_idx in [1, 3, 4]:
                    try:
                        cell.value = int(val)
                    except ValueError:
                        cell.value = val
                    cell.alignment = center_align
                elif col_idx in [2, 5]:
                    cell.value = val
                    cell.alignment = left_align
                elif col_idx in [9, 10, 11, 12, 15, 16]: # percentages
                    try:
                        float_val = float(str(val).replace("%", "")) / 100.0
                        cell.value = float_val
                    except ValueError:
                        cell.value = val
                    cell.number_format = '0.000%'
                    cell.alignment = right_align
                else: # counts and totals
                    try:
                        cell.value = int(val)
                    except ValueError:
                        cell.value = val
                    cell.number_format = '#,##0'
                    cell.alignment = right_align
                    
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        wb.save(xlsx_file)
        print(f"Successfully generated native XLSX file: {xlsx_file}")
    except ImportError:
        print("openpyxl is not installed. Native XLSX file was not generated, but the styled XLS and CSV files are ready.")

def generate_styled_xls(filename, headers, rows):
    """
    Generates an XML/HTML table spreadsheet saved as .xls.
    This file opens perfectly in Excel and contains formatting.
    """
    html = """<html xmlns:o="urn:schemas-microsoft-com:office:office"
xmlns:x="urn:schemas-microsoft-com:office:excel"
xmlns="http://www.w3.org/TR/REC-html40">
<head>
<meta http-equiv="content-type" content="text/html; charset=UTF-8">
<!--[if gte mso 9]><xml>
 <x:ExcelWorkbook>
  <x:ExcelWorksheets>
   <x:ExcelWorksheet>
    <x:Name>Actualizaciones</x:Name>
    <x:WorksheetOptions>
     <x:DisplayGridlines/>
    </x:WorksheetOptions>
   </x:ExcelWorksheet>
  </x:ExcelWorksheets>
 </x:ExcelWorkbook>
</xml><![endif]-->
<style>
  table { border-collapse: collapse; font-family: Calibri, sans-serif; font-size: 10pt; }
  th { background-color: #1F497D; color: white; font-weight: bold; border: 1px solid #D3D3D3; padding: 6px; text-align: center; }
  td { border: 1px solid #D3D3D3; padding: 5px; }
  .region { background-color: #E2EFDA; } /* light green */
  .extranjero { background-color: #FCE4D6; } /* light orange */
  .even { background-color: #F9F9F9; }
  .num { mso-number-format: "#,##0"; text-align: right; }
  .pct { mso-number-format: "0\.000%"; text-align: right; }
  .center { text-align: center; }
</style>
</head>
<body>
<table>
  <tr>
"""
    for h in headers:
        html += f"    <th>{h}</th>\n"
    html += "  </tr>\n"
    
    for r_idx, row in enumerate(rows):
        is_even = r_idx % 2 == 0
        ambito = row[2]
        
        row_class = "even" if is_even else "odd"
        if ambito == "Región":
            row_class += " region"
        else:
            row_class += " extranjero"
            
        html += f"  <tr class='{row_class}'>\n"
        for col_idx, val in enumerate(row):
            # Check alignments and formatting classes
            if col_idx in [0, 2, 3]: # update_num, ambito, ubigeo
                html += f"    <td class='center'>{val}</td>\n"
            elif col_idx in [1, 4]: # datetime, name
                html += f"    <td>{val}</td>\n"
            elif col_idx in [8, 9, 10, 11, 14, 15]: # percentages
                try:
                    numeric_val = float(str(val).replace("%", "")) / 100.0
                    html += f"    <td class='pct'>{numeric_val}</td>\n"
                except ValueError:
                    html += f"    <td class='pct'>{val}</td>\n"
            else: # counts and totals
                html += f"    <td class='num'>{val}</td>\n"
        html += "  </tr>\n"
        
    html += """</table>
</body>
</html>"""
    
    with open(filename, "w", encoding="utf-8") as f:
        f.write(html)

if __name__ == "__main__":
    main()
